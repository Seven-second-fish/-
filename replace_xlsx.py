import os
import sys
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def replace_in_workbook(file_path, old_string, new_string):
    try:
        workbook = load_workbook(filename=file_path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = cell.value.replace(old_string, new_string)
        workbook.save(file_path)
        print(f"Processed: {file_path}")
    except InvalidFileException as e:
        print(f"Skipped invalid file: {file_path} ({e})")
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")

def replace_in_xlsx_files(root_dir, old_string, new_string):
    for dirpath, dirnames, filenames in os.walk(root_dir):
        for filename in filenames:
            if filename.endswith('.xlsx') and not filename.startswith('.') and not filename.startswith('~'):
                file_path = os.path.join(dirpath, filename)
                replace_in_workbook(file_path, old_string, new_string)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py <old_string> <new_string>")
        sys.exit(1)

    folder_path = os.getcwd()  # 获取当前脚本所在文件夹的路径
    old_string = sys.argv[1]
    new_string = sys.argv[2]

    replace_in_xlsx_files(folder_path, old_string, new_string)

